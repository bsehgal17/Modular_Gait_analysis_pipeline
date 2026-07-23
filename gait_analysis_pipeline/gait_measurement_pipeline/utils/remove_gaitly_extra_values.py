import sys
from pathlib import Path
from openpyxl import load_workbook

"""
Processes all Excel files in a folder (and subfolders).
Applies specific cell-deletion/shift operations to temporal_combined and spatial_combined sheets.
"""

# ── SET YOUR FOLDER PATH HERE ─────────────────────────────────────────────────
FOLDER_PATH = (
    r"C:\Users\BhavyaSehgal\Downloads\Video_pre_preocessing\excel_outputs\rtmw"
)
# ─────────────────────────────────────────────────────────────────────────────


def get_non_empty_rows(sheet, col_idx):
    """Return row numbers containing non-empty values."""
    rows = []
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=col_idx).value is not None:
            rows.append(row)
    return rows


def trim_column_to_length(sheet, col_idx, target_length):
    """Remove values from end until column contains target_length values."""
    rows = get_non_empty_rows(sheet, col_idx)

    while len(rows) > target_length:
        sheet.cell(row=rows[-1], column=col_idx).value = None
        rows.pop()


def equalize_column_lengths(sheet, col1_name, col2_name):
    """
    Compare two columns and trim the longer one from the end.
    """
    col1 = find_column(sheet, col1_name)
    col2 = find_column(sheet, col2_name)

    if col1 is None or col2 is None:
        return

    rows1 = get_non_empty_rows(sheet, col1)
    rows2 = get_non_empty_rows(sheet, col2)

    len1 = len(rows1)
    len2 = len(rows2)

    target_length = min(len1, len2)

    if len1 > target_length:
        trim_column_to_length(sheet, col1, target_length)

    if len2 > target_length:
        trim_column_to_length(sheet, col2, target_length)

    print(f"  Matched {col1_name} ({len1}) vs {col2_name} ({len2}) -> {target_length}")


def get_column_values(sheet, col_letter):
    """Return list of (row_index, value) for all non-empty cells in a column (skipping header row 1)."""
    col_cells = []
    for row in sheet.iter_rows(
        min_row=2,
        min_col=sheet[col_letter + "1"].column,
        max_col=sheet[col_letter + "1"].column,
    ):
        for cell in row:
            col_cells.append(cell)
    return col_cells


def find_column(sheet, col_name):
    """Find column letter by header name (row 1). Returns None if not found."""
    for cell in sheet[1]:
        if cell.value and str(cell.value).strip().lower() == col_name.strip().lower():
            return cell.column
    return None


def shift_cells_up(sheet, col_idx, start_row, end_row):
    """Shift cells up from start_row to end_row in a column (delete start_row, shift up)."""
    for row in range(start_row, end_row):
        sheet.cell(row=row, column=col_idx).value = sheet.cell(
            row=row + 1, column=col_idx
        ).value
    sheet.cell(row=end_row, column=col_idx).value = None


def remove_first_value(sheet, col_idx):
    """Remove the first data value (row 2) and shift remaining cells up."""
    last_row = sheet.max_row
    shift_cells_up(sheet, col_idx, 2, last_row)


def remove_last_value(sheet, col_idx):
    """Remove the last non-empty value in a column."""
    last_row = None
    for row in range(sheet.max_row, 1, -1):
        if sheet.cell(row=row, column=col_idx).value is not None:
            last_row = row
            break
    if last_row:
        sheet.cell(row=last_row, column=col_idx).value = None


def remove_last_two_values(sheet, col_idx):
    """Remove the last two non-empty values in a column."""
    non_empty_rows = []
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=col_idx).value is not None:
            non_empty_rows.append(row)
    for row in non_empty_rows[-2:]:
        sheet.cell(row=row, column=col_idx).value = None


def remove_zeros_and_shift(sheet, col_idx):
    """Find zero values in a column, delete them, and shift remaining cells up."""
    max_row = sheet.max_row
    row = 2
    while row <= max_row:
        val = sheet.cell(row=row, column=col_idx).value
        if val == 0:
            shift_cells_up(sheet, col_idx, row, max_row)
            # Don't increment row — check the same row again (it now has the shifted value)
        else:
            row += 1


def process_temporal_combined(sheet):
    ops = {
        "pipeline_stance_times": ["remove_first", "remove_last_two"],
        "pipeline_stride_times": ["remove_first"],
        "pipeline_step_times": ["remove_first"],
        "pipeline_swing_times": ["remove_last"],
        "gaitrite_swing_times": ["remove_zeros"],
        "gaitrite_stance_times": ["remove_zeros"],
    }

    for col_name, actions in ops.items():
        col_idx = find_column(sheet, col_name)
        if col_idx is None:
            print(f"  [WARN] Column '{col_name}' not found in temporal_combined sheet.")
            continue
        for action in actions:
            if action == "remove_first":
                remove_first_value(sheet, col_idx)
            elif action == "remove_last":
                remove_last_value(sheet, col_idx)
            elif action == "remove_last_two":
                remove_last_two_values(sheet, col_idx)
            elif action == "remove_zeros":
                remove_zeros_and_shift(sheet, col_idx)
        # Match pipeline/gaitrite lengths
    equalize_column_lengths(
        sheet,
        "pipeline_step_times",
        "gaitrite_step_times",
    )

    equalize_column_lengths(
        sheet,
        "pipeline_stride_times",
        "gaitrite_stride_times",
    )

    equalize_column_lengths(
        sheet,
        "pipeline_stance_times",
        "gaitrite_stance_times",
    )

    equalize_column_lengths(
        sheet,
        "pipeline_swing_times",
        "gaitrite_swing_times",
    )


def process_spatial_combined(sheet):
    ops = {
        "pipeline_Stride_lengths": ["remove_first"],
        "pipeline_Step_lengths": ["remove_first"],
    }

    for col_name, actions in ops.items():
        col_idx = find_column(sheet, col_name)
        if col_idx is None:
            print(f"  [WARN] Column '{col_name}' not found in spatial_combined sheet.")
            continue
        for action in actions:
            if action == "remove_first":
                remove_first_value(sheet, col_idx)
    equalize_column_lengths(
        sheet,
        "pipeline_Stride_lengths",
        "gaitrite_Stride_lengths",
    )

    equalize_column_lengths(
        sheet,
        "pipeline_Step_lengths",
        "gaitrite_Step_lengths",
    )


def process_file(filepath):
    print(f"Processing: {filepath}")
    try:
        wb = load_workbook(filepath)
    except Exception as e:
        print(f"  [ERROR] Could not open file: {e}")
        return

    if "temporal_combined" in wb.sheetnames:
        process_temporal_combined(wb["temporal_combined"])
    else:
        print("  [WARN] Sheet 'temporal_combined' not found.")

    if "spatial_combined" in wb.sheetnames:
        process_spatial_combined(wb["spatial_combined"])
    else:
        print("  [WARN] Sheet 'spatial_combined' not found.")

    try:
        wb.save(filepath)
        print(f"  Saved.")
    except Exception as e:
        print(f"  [ERROR] Could not save file: {e}")


def process_folder(folder_path):
    folder = Path(folder_path)
    if not folder.is_dir():
        print(f"Error: '{folder_path}' is not a valid directory.")
        sys.exit(1)

    xlsx_files = list(folder.rglob("*.xlsx")) + list(folder.rglob("*.xlsm"))
    if not xlsx_files:
        print("No Excel files found.")
        return

    print(f"Found {len(xlsx_files)} Excel file(s).\n")
    for f in xlsx_files:
        process_file(f)

    print("\nDone.")
