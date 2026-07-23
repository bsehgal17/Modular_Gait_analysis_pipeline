from dataclasses import asdict
from pathlib import Path
import json
from typing import List, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from gait_measurement_pipeline.gait_evaluation.json_loader import JSONResultsLoader
from gait_measurement_pipeline.gait_evaluation.gait_comparison import (
    GaitMetricsComparator,
)
from gait_measurement_pipeline.utils.remove_gaitly_extra_values import process_folder


class GaitEvaluationRunner:
    """
    End-to-end gait evaluation runner.

    Responsibilities:
    -----------------
    1. Load JSON results from:
        - Pipeline output
        - Reference GaitRite output
    2. Export raw per-pass Excel files for manual inspection/editing
    3. Compare temporal, spatial, and derived gait metrics
    4. Save final comparison results as a JSON file
    """

    def __init__(
        self,
        pipeline_folder: Path,
        gaitrite_folder: Path,
        output_file: Path,
    ) -> None:
        """
        Initialize the evaluation runner.

        Args:
            pipeline_folder (Path): Directory containing pipeline JSON results.
            gaitrite_folder (Path): Directory containing GaitRite JSON results.
            output_file (Path): File path where comparison JSON will be saved.
        """
        self.pipeline_folder: Path = Path(pipeline_folder)
        self.gaitrite_folder: Path = Path(gaitrite_folder)
        self.output_file: Path = Path(output_file)

    # -------------------------
    # Main execution method
    # -------------------------
    def run(self) -> List[Any]:
        """
        Execute the full evaluation pipeline.

        Workflow:
            1. Load pipeline results
            2. Load GaitRite reference results
            3. Export raw per-pass Excel files
            4. Compare metrics using Excel inputs
            5. Save comparison results to JSON

        Returns:
            List[Any]: List of comparison result objects
                       (typically custom objects from comparator).
        """

        # -------------------------
        # Step 1: Load pipeline results
        # -------------------------
        print("Loading pipeline results...")
        pipeline_results: List[Any] = JSONResultsLoader(self.pipeline_folder).load()
        print(f"Loaded {len(pipeline_results)} pipeline passes.")

        # -------------------------
        # Step 2: Load GaitRite results
        # -------------------------
        print("Loading GaitRite results...")
        gaitrite_results: List[Any] = JSONResultsLoader(self.gaitrite_folder).load()
        print(f"Loaded {len(gaitrite_results)} GaitRite passes.")

        # -------------------------
        # Step 3: Initialize comparator
        # -------------------------
        print("Initializing comparator...")
        comparator: GaitMetricsComparator = GaitMetricsComparator(
            pipeline_results,
            gaitrite_results,
        )

        # Directory where Excel files will be exported
        excel_path: Path = Path(
            r"C:\Users\BhavyaSehgal\Downloads\Video_pre_preocessing\excel_outputs\rtmw"
        )

        # -------------------------
        # Step 4: Export raw Excel per pass
        # -------------------------
        print("Exporting raw per-pass Excel files...")
        comparator.export_raw_per_pass(output_dir=excel_path)

        # remove pipeline extra steps and match gaitrite steps
        process_folder(excel_path)

        # -------------------------
        # Step 5: Perform comparison using Excel inputs
        # -------------------------
        print("Running metric comparison...")
        comparison_results: List[Any] = comparator.compare(excel_root=excel_path)

        # -------------------------
        # Step 6: Save results to JSON
        # -------------------------
        self._save(comparison_results)
        print(f"Saved comparison results to {self.output_file}")

        return comparison_results

    # -------------------------
    # Save helper method
    # -------------------------
    def _save(self, results: List[Any]) -> None:
        self.output_file.parent.mkdir(parents=True, exist_ok=True)

        raw = [asdict(r) for r in results]

        wb = openpyxl.Workbook()

        SUB_STATS = [
            "mae",
            "rmse",
            "percent_error",
            "std",
            "bias",
            "r",
            "icc",
            "loa_lower",
            "loa_upper",
        ]

        GROUPS = [
            (
                "temporal_stats",
                ["stance_times", "stride_times", "step_times", "swing_times"],
            ),
            ("spatial_stats", ["Stride_lengths", "Step_lengths"]),
        ]

        # Styles
        black_font = Font(name="Arial", bold=True, color="000000", size=13)
        subheader_font = Font(name="Arial", bold=True, color="000000", size=10)
        cell_font = Font(name="Arial", size=10, color="000000")
        participant_font = Font(name="Arial", bold=True, size=10, color="000000")
        white_fill = PatternFill("solid", start_color="FFFFFF")
        alt_fill = PatternFill("solid", start_color="F2F2F2")
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style(cell, font=None, fill=None, align=None, num_fmt=None):
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if align:
                cell.alignment = align
            if num_fmt:
                cell.number_format = num_fmt
            cell.border = border

        for group_key, metrics in GROUPS:
            for metric in metrics:
                sheet_name = metric.replace("_", " ").title()
                ws = wb.create_sheet(sheet_name)

                # Row 1: metric title banner
                ws.merge_cells(
                    start_row=1,
                    start_column=1,
                    end_row=1,
                    end_column=len(SUB_STATS) + 1,
                )
                cell = ws.cell(1, 1, sheet_name)
                style(
                    cell,
                    font=Font(name="Arial", bold=True, color="000000", size=13),
                    fill=white_fill,
                    align=center,
                )

                # Row 2: column headers
                cell = ws.cell(2, 1, "Participant")
                style(
                    cell,
                    font=Font(name="Arial", bold=True, color="000000", size=10),
                    fill=white_fill,
                    align=center,
                )
                for i, s in enumerate(SUB_STATS):
                    cell = ws.cell(2, i + 2, s.upper().replace("_", " "))
                    style(
                        cell,
                        font=Font(name="Arial", bold=True, color="000000", size=10),
                        fill=white_fill,
                        align=center,
                    )

                # Rows 3+: one participant per row
                for row_i, entry in enumerate(raw):
                    row = 3 + row_i
                    fill = alt_fill if row_i % 2 == 0 else white_fill
                    cell = ws.cell(row, 1, entry["participant_id"])
                    style(cell, font=participant_font, fill=fill, align=left)

                    stats = entry[group_key].get(metric, {})
                    for i, s in enumerate(SUB_STATS):
                        val = stats.get(s)
                        cell = ws.cell(
                            row, i + 2, round(val, 4) if val is not None else ""
                        )
                        fmt = '0.00"%"' if s == "percent_error" else "0.0000"
                        style(
                            cell, font=cell_font, fill=fill, align=center, num_fmt=fmt
                        )

                # Summary row
                summary_row = 3 + len(raw)
                cell = ws.cell(summary_row, 1, "Average")
                style(
                    cell,
                    font=Font(name="Arial", bold=True, size=10, color="000000"),
                    fill=white_fill,
                    align=left,
                )
                for i in range(len(SUB_STATS)):
                    col_letter = get_column_letter(i + 2)
                    cell = ws.cell(
                        summary_row,
                        i + 2,
                        f"=AVERAGE({col_letter}3:{col_letter}{summary_row - 1})",
                    )
                    fmt = '0.00"%"' if SUB_STATS[i] == "percent_error" else "0.0000"
                    style(
                        cell,
                        font=Font(name="Arial", bold=True, size=10, color="000000"),
                        fill=white_fill,
                        align=center,
                        num_fmt=fmt,
                    )

                # Column widths
                ws.column_dimensions["A"].width = 24
                for col in range(2, len(SUB_STATS) + 2):
                    ws.column_dimensions[get_column_letter(col)].width = 20
                ws.row_dimensions[1].height = 24
                ws.row_dimensions[2].height = 20
                ws.freeze_panes = "B3"

        del wb["Sheet"]
        wb.save(self.output_file.with_suffix(".xlsx"))
